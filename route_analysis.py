from __future__ import annotations

import colorsys
import gzip
import math
import pickle
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from scipy.optimize import linear_sum_assignment
from sklearn.metrics import adjusted_rand_score, normalized_mutual_info_score


EARTH_RADIUS_KM = 6371.0088
REQUIRED_COLUMNS = {"LOTE", "UL", "INSTALACAO", "LATITUDE", "LONGITUDE"}
DEFAULT_CURRENT_FILE = Path("Rotas_Atuais_VV_Completo.xlsx")
DEFAULT_OPTIMIZED_FILE = Path("Rota_Pronta_VV_Consolidado.xlsx")
DEFAULT_OUTPUT_FILE = Path("outputs/Comparativo_Rotas_VV.xlsx")
DEFAULT_DASHBOARD_SNAPSHOT = Path("outputs/dashboard_snapshot.pkl.gz")
DASHBOARD_SNAPSHOT_VERSION = 2


@dataclass
class SourceLoad:
    data: pd.DataFrame
    quality_summary: pd.DataFrame
    quality_issues: pd.DataFrame
    source_info: dict[str, Any]


def _clean_identifier(value: Any) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        candidate = text[:-2]
        if candidate.lstrip("-").isdigit():
            return candidate
    return text or None


def _numeric_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce").astype(float)
    cleaned = series.astype("string").str.strip().str.replace(",", ".", regex=False)
    return pd.to_numeric(cleaned, errors="coerce").astype(float)


def normalize_coordinate(series: pd.Series, limit: float) -> tuple[pd.Series, pd.Series]:
    values = _numeric_series(series)
    steps = pd.Series(np.zeros(len(values), dtype=np.int16), index=values.index)
    for _ in range(12):
        mask = values.abs() > limit
        if not bool(mask.any()):
            break
        values.loc[mask] = values.loc[mask] / 10.0
        steps.loc[mask] = steps.loc[mask] + 1
    return values, steps


def load_source(path: str | Path, scenario: str) -> SourceLoad:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    raw = pd.read_excel(path)
    raw.columns = [str(column).strip().upper() for column in raw.columns]
    missing = REQUIRED_COLUMNS.difference(raw.columns)
    if missing:
        raise ValueError(
            f"{path.name} não possui as colunas obrigatórias: {', '.join(sorted(missing))}"
        )

    frame = raw[["LOTE", "UL", "INSTALACAO", "LATITUDE", "LONGITUDE"]].copy()
    frame["lot"] = pd.to_numeric(frame["LOTE"], errors="coerce")
    frame["route"] = frame["UL"].map(_clean_identifier)
    frame["installation"] = frame["INSTALACAO"].map(_clean_identifier)
    frame["latitude"], lat_steps = normalize_coordinate(frame["LATITUDE"], 90.0)
    frame["longitude"], lon_steps = normalize_coordinate(frame["LONGITUDE"], 180.0)

    valid_key = frame["lot"].notna() & frame["route"].notna() & frame["installation"].notna()
    valid_coordinate = (
        frame["latitude"].between(-90, 90, inclusive="both")
        & frame["longitude"].between(-180, 180, inclusive="both")
    )
    duplicate_key = frame.duplicated(["lot", "installation"], keep="first") & valid_key

    usable = frame.loc[valid_key & valid_coordinate & ~duplicate_key].copy()
    usable["lot"] = usable["lot"].astype(int)
    usable["scenario"] = scenario
    usable = usable[
        ["scenario", "lot", "route", "installation", "latitude", "longitude"]
    ].reset_index(drop=True)

    issue_rows: list[dict[str, Any]] = []
    invalid_rows = frame.loc[~valid_key | ~valid_coordinate].head(1000)
    for _, row in invalid_rows.iterrows():
        issue_rows.append(
            {
                "scenario": scenario,
                "issue": "Registro inválido",
                "lot": row.get("lot"),
                "installation": row.get("installation"),
                "route": row.get("route"),
                "detail": "Chave obrigatória ou coordenada inválida",
            }
        )

    duplicated_rows = frame.loc[duplicate_key].head(1000)
    for _, row in duplicated_rows.iterrows():
        issue_rows.append(
            {
                "scenario": scenario,
                "issue": "Duplicidade lote-instalação",
                "lot": row.get("lot"),
                "installation": row.get("installation"),
                "route": row.get("route"),
                "detail": "Mantida somente a primeira ocorrência para os cálculos",
            }
        )

    cross_lot = (
        usable.groupby("installation", sort=False)["lot"]
        .nunique()
        .loc[lambda values: values > 1]
    )
    for installation, count in cross_lot.head(1000).items():
        lots = usable.loc[usable["installation"].eq(installation), "lot"].tolist()
        issue_rows.append(
            {
                "scenario": scenario,
                "issue": "Instalação em múltiplos lotes",
                "lot": ", ".join(map(str, sorted(lots))),
                "installation": installation,
                "route": None,
                "detail": f"Instalação encontrada em {int(count)} lotes",
            }
        )

    quality_summary = pd.DataFrame(
        [
            {
                "scenario": scenario,
                "metric": "Linhas no arquivo",
                "value": int(len(raw)),
                "note": path.name,
            },
            {
                "scenario": scenario,
                "metric": "Linhas usadas nos cálculos",
                "value": int(len(usable)),
                "note": "Após validação e deduplicação por lote-instalação",
            },
            {
                "scenario": scenario,
                "metric": "Coordenadas reescaladas",
                "value": int(((lat_steps > 0) | (lon_steps > 0)).sum()),
                "note": "Casas decimais restauradas automaticamente",
            },
            {
                "scenario": scenario,
                "metric": "Registros inválidos",
                "value": int((~valid_key | ~valid_coordinate).sum()),
                "note": "Não entram nos indicadores espaciais",
            },
            {
                "scenario": scenario,
                "metric": "Duplicidades lote-instalação",
                "value": int(duplicate_key.sum()),
                "note": "Mantida a primeira ocorrência",
            },
            {
                "scenario": scenario,
                "metric": "Instalações em múltiplos lotes",
                "value": int(len(cross_lot)),
                "note": "Mantidas, mas sinalizadas para conferência",
            },
        ]
    )
    quality_issues = pd.DataFrame(
        issue_rows,
        columns=["scenario", "issue", "lot", "installation", "route", "detail"],
    )
    stat = path.stat()
    source_info = {
        "scenario": scenario,
        "path": str(path.resolve()),
        "file_name": path.name,
        "modified_at": datetime.fromtimestamp(stat.st_mtime),
        "size_bytes": int(stat.st_size),
        "rows": int(len(raw)),
        "usable_rows": int(len(usable)),
    }
    return SourceLoad(usable, quality_summary, quality_issues, source_info)


def haversine_to_point(
    latitudes: np.ndarray,
    longitudes: np.ndarray,
    point_latitude: float,
    point_longitude: float,
) -> np.ndarray:
    lat1 = np.radians(np.asarray(latitudes, dtype=float))
    lon1 = np.radians(np.asarray(longitudes, dtype=float))
    lat2 = math.radians(float(point_latitude))
    lon2 = math.radians(float(point_longitude))
    dlat = lat1 - lat2
    dlon = lon1 - lon2
    a = np.sin(dlat / 2.0) ** 2 + np.cos(lat1) * math.cos(lat2) * np.sin(dlon / 2.0) ** 2
    return EARTH_RADIUS_KM * 2.0 * np.arcsin(np.sqrt(np.clip(a, 0.0, 1.0)))


def haversine_matrix(latitudes: np.ndarray, longitudes: np.ndarray) -> np.ndarray:
    lat = np.radians(np.asarray(latitudes, dtype=float))
    lon = np.radians(np.asarray(longitudes, dtype=float))
    dlat = lat[:, None] - lat[None, :]
    dlon = lon[:, None] - lon[None, :]
    a = np.sin(dlat / 2.0) ** 2 + (
        np.cos(lat[:, None]) * np.cos(lat[None, :]) * np.sin(dlon / 2.0) ** 2
    )
    return EARTH_RADIUS_KM * 2.0 * np.arcsin(np.sqrt(np.clip(a, 0.0, 1.0)))


def _stats(values: np.ndarray, prefix: str) -> dict[str, float]:
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        return {
            f"{prefix}_mean_km": np.nan,
            f"{prefix}_median_km": np.nan,
            f"{prefix}_max_km": np.nan,
        }
    return {
        f"{prefix}_mean_km": float(np.mean(values)),
        f"{prefix}_median_km": float(np.median(values)),
        f"{prefix}_max_km": float(np.max(values)),
    }


def nearest_neighbor_path_total(
    distance_matrix: np.ndarray,
    start_index: int,
) -> tuple[list[int], float]:
    """Constrói um caminho aberto guloso, sempre visitando o ponto não visitado mais próximo."""
    distances = np.asarray(distance_matrix, dtype=float)
    if distances.ndim != 2 or distances.shape[0] != distances.shape[1]:
        raise ValueError("A matriz de distâncias deve ser quadrada.")
    count = int(distances.shape[0])
    if count == 0:
        return [], 0.0
    if not 0 <= int(start_index) < count:
        raise IndexError("Índice inicial fora da matriz de distâncias.")

    visited = np.zeros(count, dtype=bool)
    current = int(start_index)
    visited[current] = True
    order = [current]
    total = 0.0
    while len(order) < count:
        candidates = np.flatnonzero(~visited)
        candidate_distances = distances[current, candidates]
        next_index = int(candidates[int(np.argmin(candidate_distances))])
        segment = float(distances[current, next_index])
        if not np.isfinite(segment):
            raise ValueError("A matriz contém distância não finita entre pontos visitáveis.")
        total += segment
        visited[next_index] = True
        order.append(next_index)
        current = next_index
    return order, total


def _euclidean_matrix_km(coordinates_m: np.ndarray) -> np.ndarray:
    coordinates = np.asarray(coordinates_m, dtype=float)
    differences = coordinates[:, None, :] - coordinates[None, :, :]
    return np.sqrt(np.sum(differences * differences, axis=2)) / 1000.0


def compute_spatial_metrics(data: pd.DataFrame, scenario: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    import geopandas as gpd

    route_records: list[dict[str, Any]] = []
    lot_records: list[dict[str, Any]] = []

    for lot, lot_data in data.groupby("lot", sort=True):
        lot_centroid_lat = float(lot_data["latitude"].mean())
        lot_centroid_lon = float(lot_data["longitude"].mean())
        geographic_points = gpd.GeoDataFrame(
            index=lot_data.index,
            geometry=gpd.points_from_xy(lot_data["longitude"], lot_data["latitude"]),
            crs="EPSG:4326",
        )
        projected_crs = geographic_points.estimate_utm_crs()
        if projected_crs is None:
            raise ValueError(f"Não foi possível determinar a projeção métrica do lote {int(lot)}.")
        projected_points = geographic_points.to_crs(projected_crs)
        projected_coordinates = pd.DataFrame(
            {
                "x": projected_points.geometry.x.to_numpy(float),
                "y": projected_points.geometry.y.to_numpy(float),
            },
            index=lot_data.index,
        )
        pair_arrays: list[np.ndarray] = []
        nearest_arrays: list[np.ndarray] = []
        radius_arrays: list[np.ndarray] = []
        lot_route_indices: list[int] = []

        for route, route_data in lot_data.groupby("route", sort=True):
            latitudes = route_data["latitude"].to_numpy(float)
            longitudes = route_data["longitude"].to_numpy(float)
            count = int(len(route_data))
            centroid_lat = float(np.mean(latitudes))
            centroid_lon = float(np.mean(longitudes))
            radius = haversine_to_point(latitudes, longitudes, centroid_lat, centroid_lon)

            if count >= 2:
                matrix = haversine_matrix(latitudes, longitudes)
                pair_distances = matrix[np.triu_indices(count, 1)]
                np.fill_diagonal(matrix, np.inf)
                nearest_distances = np.min(matrix, axis=1)
                haversine_start = int(np.argmax(radius))
                _, nn_path_haversine_km = nearest_neighbor_path_total(
                    matrix, haversine_start
                )
                del matrix

                route_projected = projected_coordinates.loc[
                    route_data.index, ["x", "y"]
                ].to_numpy(float)
                projected_centroid = np.mean(route_projected, axis=0)
                projected_radius = np.sqrt(
                    np.sum((route_projected - projected_centroid) ** 2, axis=1)
                )
                projected_start = int(np.argmax(projected_radius))
                projected_matrix = _euclidean_matrix_km(route_projected)
                _, nn_path_geopandas_km = nearest_neighbor_path_total(
                    projected_matrix, projected_start
                )
                del projected_matrix
            else:
                pair_distances = np.array([], dtype=float)
                nearest_distances = np.array([], dtype=float)
                nn_path_haversine_km = 0.0
                nn_path_geopandas_km = 0.0

            record: dict[str, Any] = {
                "scenario": scenario,
                "lot": int(lot),
                "route": str(route),
                "installations": count,
                "centroid_latitude": centroid_lat,
                "centroid_longitude": centroid_lon,
                "pair_zero_pct": float(np.mean(pair_distances <= 1e-9)) if pair_distances.size else np.nan,
                "nearest_zero_pct": float(np.mean(nearest_distances <= 1e-9)) if nearest_distances.size else np.nan,
                "route_to_lot_centroid_km": float(
                    haversine_to_point(
                        np.array([centroid_lat]),
                        np.array([centroid_lon]),
                        lot_centroid_lat,
                        lot_centroid_lon,
                    )[0]
                ),
                "nn_path_haversine_km": float(nn_path_haversine_km),
                "nn_path_geopandas_km": float(nn_path_geopandas_km),
            }
            record.update(_stats(pair_distances, "pair"))
            record.update(_stats(nearest_distances, "nearest"))
            record.update(_stats(radius, "radius"))
            route_records.append(record)
            lot_route_indices.append(len(route_records) - 1)
            pair_arrays.append(pair_distances)
            nearest_arrays.append(nearest_distances)
            radius_arrays.append(radius)

        centroid_latitudes = np.array(
            [route_records[index]["centroid_latitude"] for index in lot_route_indices], dtype=float
        )
        centroid_longitudes = np.array(
            [route_records[index]["centroid_longitude"] for index in lot_route_indices], dtype=float
        )
        if len(lot_route_indices) >= 2:
            centroid_matrix = haversine_matrix(centroid_latitudes, centroid_longitudes)
            np.fill_diagonal(centroid_matrix, np.inf)
            nearest_centroids = np.min(centroid_matrix, axis=1)
        else:
            nearest_centroids = np.full(len(lot_route_indices), np.nan)

        for local_index, record_index in enumerate(lot_route_indices):
            nearest_centroid = float(nearest_centroids[local_index])
            radius_mean = route_records[record_index]["radius_mean_km"]
            route_records[record_index]["nearest_route_centroid_km"] = nearest_centroid
            route_records[record_index]["separation_index"] = (
                nearest_centroid / radius_mean
                if np.isfinite(nearest_centroid) and np.isfinite(radius_mean) and radius_mean > 0
                else np.nan
            )

        all_pairs = np.concatenate(pair_arrays) if pair_arrays else np.array([], dtype=float)
        all_nearest = np.concatenate(nearest_arrays) if nearest_arrays else np.array([], dtype=float)
        all_radius = np.concatenate(radius_arrays) if radius_arrays else np.array([], dtype=float)
        route_sizes = lot_data.groupby("route", sort=False).size().to_numpy(float)
        route_to_lot = np.array(
            [route_records[index]["route_to_lot_centroid_km"] for index in lot_route_indices],
            dtype=float,
        )
        separation = np.array(
            [route_records[index]["separation_index"] for index in lot_route_indices], dtype=float
        )
        nn_path_haversine = np.array(
            [route_records[index]["nn_path_haversine_km"] for index in lot_route_indices],
            dtype=float,
        )
        nn_path_geopandas = np.array(
            [route_records[index]["nn_path_geopandas_km"] for index in lot_route_indices],
            dtype=float,
        )

        lot_record: dict[str, Any] = {
            "scenario": scenario,
            "lot": int(lot),
            "installations": int(len(lot_data)),
            "routes": int(len(lot_route_indices)),
            "lot_centroid_latitude": lot_centroid_lat,
            "lot_centroid_longitude": lot_centroid_lon,
            "pair_zero_pct": float(np.mean(all_pairs <= 1e-9)) if all_pairs.size else np.nan,
            "nearest_zero_pct": float(np.mean(all_nearest <= 1e-9)) if all_nearest.size else np.nan,
            "route_size_min": float(np.min(route_sizes)),
            "route_size_mean": float(np.mean(route_sizes)),
            "route_size_median": float(np.median(route_sizes)),
            "route_size_max": float(np.max(route_sizes)),
            "route_size_std": float(np.std(route_sizes, ddof=1)) if route_sizes.size > 1 else 0.0,
            "route_size_cv": (
                float(np.std(route_sizes, ddof=1) / np.mean(route_sizes))
                if route_sizes.size > 1 and np.mean(route_sizes) > 0
                else 0.0
            ),
            "separation_index_mean": (
                float(np.nanmean(separation)) if np.isfinite(separation).any() else np.nan
            ),
            "separation_index_median": (
                float(np.nanmedian(separation)) if np.isfinite(separation).any() else np.nan
            ),
            "nn_path_haversine_km": float(np.sum(nn_path_haversine)),
            "nn_path_geopandas_km": float(np.sum(nn_path_geopandas)),
        }
        lot_record.update(_stats(all_pairs, "pair"))
        lot_record.update(_stats(all_nearest, "nearest"))
        lot_record.update(_stats(all_radius, "radius"))
        lot_record.update(_stats(route_to_lot, "route_to_lot_centroid"))
        lot_records.append(lot_record)

    return pd.DataFrame(route_records), pd.DataFrame(lot_records)


def _choose_status(has_current: bool, has_optimized: bool) -> str:
    if has_current and has_optimized:
        return "Comparável"
    if has_current:
        return "Aguardando otimização"
    return "Sem base atual"


def compute_group_comparison(
    current: pd.DataFrame,
    optimized: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    current_lots = set(current["lot"].unique())
    optimized_lots = set(optimized["lot"].unique())
    all_lots = sorted(current_lots | optimized_lots)

    coverage_records: list[dict[str, Any]] = []
    transition_frames: list[pd.DataFrame] = []
    stability_records: list[dict[str, Any]] = []
    fragmentation_records: list[dict[str, Any]] = []

    comparable_lots = sorted(current_lots & optimized_lots)
    current_match = current.loc[current["lot"].isin(comparable_lots), [
        "lot", "installation", "route", "latitude", "longitude"
    ]].rename(columns={
        "route": "current_route",
        "latitude": "current_latitude",
        "longitude": "current_longitude",
    })
    optimized_match = optimized.loc[optimized["lot"].isin(comparable_lots), [
        "lot", "installation", "route", "latitude", "longitude"
    ]].rename(columns={
        "route": "optimized_route",
        "latitude": "optimized_latitude",
        "longitude": "optimized_longitude",
    })
    installation_match = current_match.merge(
        optimized_match,
        on=["lot", "installation"],
        how="outer",
        validate="one_to_one",
    )
    installation_match["status"] = np.select(
        [
            installation_match["current_route"].notna() & installation_match["optimized_route"].notna(),
            installation_match["current_route"].notna(),
        ],
        ["Comum", "Somente atual"],
        default="Somente otimizada",
    )
    common_coordinate = installation_match["status"].eq("Comum")
    installation_match["coordinate_difference_m"] = np.nan
    if bool(common_coordinate.any()):
        lat1 = installation_match.loc[common_coordinate, "current_latitude"].to_numpy(float)
        lon1 = installation_match.loc[common_coordinate, "current_longitude"].to_numpy(float)
        lat2 = installation_match.loc[common_coordinate, "optimized_latitude"].to_numpy(float)
        lon2 = installation_match.loc[common_coordinate, "optimized_longitude"].to_numpy(float)
        lat1r, lon1r, lat2r, lon2r = map(np.radians, (lat1, lon1, lat2, lon2))
        delta_lat = lat2r - lat1r
        delta_lon = lon2r - lon1r
        a = np.sin(delta_lat / 2.0) ** 2 + np.cos(lat1r) * np.cos(lat2r) * np.sin(delta_lon / 2.0) ** 2
        coordinate_diff = EARTH_RADIUS_KM * 2000.0 * np.arcsin(np.sqrt(np.clip(a, 0.0, 1.0)))
        installation_match.loc[common_coordinate, "coordinate_difference_m"] = coordinate_diff

    for lot in all_lots:
        current_lot = current.loc[current["lot"].eq(lot)]
        optimized_lot = optimized.loc[optimized["lot"].eq(lot)]
        current_ids = set(current_lot["installation"])
        optimized_ids = set(optimized_lot["installation"])
        common_ids = current_ids & optimized_ids
        coverage_records.append(
            {
                "lot": int(lot),
                "status": _choose_status(bool(len(current_lot)), bool(len(optimized_lot))),
                "current_installations": int(len(current_ids)),
                "optimized_installations": int(len(optimized_ids)),
                "common_installations": int(len(common_ids)),
                "current_only_installations": int(len(current_ids - optimized_ids)),
                "optimized_only_installations": int(len(optimized_ids - current_ids)),
                "current_routes": int(current_lot["route"].nunique()),
                "optimized_routes": int(optimized_lot["route"].nunique()),
                "current_coverage_pct": (
                    len(common_ids) / len(current_ids) if current_ids else np.nan
                ),
            }
        )
        if not len(current_lot) or not len(optimized_lot):
            continue

        common = current_lot[["lot", "installation", "route"]].merge(
            optimized_lot[["lot", "installation", "route"]],
            on=["lot", "installation"],
            suffixes=("_current", "_optimized"),
            validate="one_to_one",
        )
        if common.empty:
            continue

        transition = (
            common.groupby(["route_current", "route_optimized"], sort=True)
            .size()
            .rename("installations")
            .reset_index()
            .rename(columns={
                "route_current": "current_route",
                "route_optimized": "optimized_route",
            })
        )
        current_sizes = common.groupby("route_current").size().rename("current_route_installations")
        optimized_sizes = common.groupby("route_optimized").size().rename("optimized_route_installations")
        transition = transition.merge(
            current_sizes, left_on="current_route", right_index=True
        ).merge(optimized_sizes, left_on="optimized_route", right_index=True)
        transition["lot"] = int(lot)
        transition["share_current_route"] = (
            transition["installations"] / transition["current_route_installations"]
        )
        transition["share_optimized_route"] = (
            transition["installations"] / transition["optimized_route_installations"]
        )
        transition["retained_pairs_in_cell"] = (
            transition["installations"] * (transition["installations"] - 1) // 2
        ).astype(np.int64)
        transition_frames.append(transition)

        old_pairs = int(((current_sizes * (current_sizes - 1)) // 2).sum())
        new_pairs = int(((optimized_sizes * (optimized_sizes - 1)) // 2).sum())
        retained_pairs = int(transition["retained_pairs_in_cell"].sum())
        separated_pairs = old_pairs - retained_pairs
        newly_grouped_pairs = new_pairs - retained_pairs
        union_pairs = old_pairs + new_pairs - retained_pairs
        stability_records.append(
            {
                "lot": int(lot),
                "common_installations": int(len(common)),
                "current_pairs": old_pairs,
                "optimized_pairs": new_pairs,
                "retained_pairs": retained_pairs,
                "separated_pairs": separated_pairs,
                "newly_grouped_pairs": newly_grouped_pairs,
                "pair_retention": retained_pairs / old_pairs if old_pairs else np.nan,
                "pair_precision": retained_pairs / new_pairs if new_pairs else np.nan,
                "pair_jaccard": retained_pairs / union_pairs if union_pairs else np.nan,
                "ari": float(adjusted_rand_score(common["route_current"], common["route_optimized"])),
                "nmi": float(
                    normalized_mutual_info_score(common["route_current"], common["route_optimized"])
                ),
            }
        )

        for current_route, group in transition.groupby("current_route", sort=True):
            route_total = int(group["current_route_installations"].iloc[0])
            dominant_row = group.loc[group["installations"].idxmax()]
            route_pairs = route_total * (route_total - 1) // 2
            retained_in_route = int(group["retained_pairs_in_cell"].sum())
            probabilities = group["installations"].to_numpy(float) / route_total
            fragmentation_records.append(
                {
                    "lot": int(lot),
                    "current_route": str(current_route),
                    "common_installations": route_total,
                    "optimized_routes_received": int(group["optimized_route"].nunique()),
                    "dominant_optimized_route": str(dominant_row["optimized_route"]),
                    "dominant_installations": int(dominant_row["installations"]),
                    "dominant_share": float(dominant_row["installations"] / route_total),
                    "fragmentation_index": float(1.0 - np.sum(probabilities**2)),
                    "separated_pairs": route_pairs - retained_in_route,
                    "separated_pairs_pct": (
                        (route_pairs - retained_in_route) / route_pairs if route_pairs else 0.0
                    ),
                }
            )

    transitions = (
        pd.concat(transition_frames, ignore_index=True)
        if transition_frames
        else pd.DataFrame(
            columns=[
                "current_route", "optimized_route", "installations",
                "current_route_installations", "optimized_route_installations", "lot",
                "share_current_route", "share_optimized_route", "retained_pairs_in_cell",
            ]
        )
    )
    return {
        "coverage": pd.DataFrame(coverage_records),
        "transitions": transitions,
        "stability": pd.DataFrame(stability_records),
        "fragmentation": pd.DataFrame(fragmentation_records),
        "installation_match": installation_match.sort_values(["lot", "installation"]).reset_index(drop=True),
    }


def run_analysis(
    current_file: str | Path = DEFAULT_CURRENT_FILE,
    optimized_file: str | Path = DEFAULT_OPTIMIZED_FILE,
) -> dict[str, Any]:
    current_source = load_source(current_file, "Atual")
    optimized_source = load_source(optimized_file, "Otimizada")
    current_routes, current_lots = compute_spatial_metrics(current_source.data, "Atual")
    optimized_routes, optimized_lots = compute_spatial_metrics(optimized_source.data, "Otimizada")
    group_comparison = compute_group_comparison(current_source.data, optimized_source.data)

    return {
        "current": current_source.data,
        "optimized": optimized_source.data,
        "route_metrics": pd.concat([current_routes, optimized_routes], ignore_index=True),
        "lot_metrics": pd.concat([current_lots, optimized_lots], ignore_index=True),
        "coverage": group_comparison["coverage"],
        "transitions": group_comparison["transitions"],
        "stability": group_comparison["stability"],
        "fragmentation": group_comparison["fragmentation"],
        "installation_match": group_comparison["installation_match"],
        "quality_summary": pd.concat(
            [current_source.quality_summary, optimized_source.quality_summary], ignore_index=True
        ),
        "quality_issues": pd.concat(
            [current_source.quality_issues, optimized_source.quality_issues], ignore_index=True
        ),
        "source_info": [current_source.source_info, optimized_source.source_info],
        "generated_at": datetime.now(),
    }


def build_lot_comparison(analysis: dict[str, Any]) -> pd.DataFrame:
    coverage = analysis["coverage"].copy()
    lots = analysis["lot_metrics"].copy()
    current = lots.loc[lots["scenario"].eq("Atual")].drop(
        columns=["scenario", "installations", "routes"]
    )
    current = current.rename(
        columns={column: f"current_{column}" for column in current.columns if column != "lot"}
    )
    optimized = lots.loc[lots["scenario"].eq("Otimizada")].drop(
        columns=["scenario", "installations", "routes"]
    )
    optimized = optimized.rename(
        columns={column: f"optimized_{column}" for column in optimized.columns if column != "lot"}
    )
    comparison = coverage.merge(current, on="lot", how="left").merge(
        optimized, on="lot", how="left"
    )
    stability = analysis["stability"].copy()
    comparison = comparison.merge(stability, on="lot", how="left", suffixes=("", "_stability"))
    comparison["route_delta"] = comparison["optimized_routes"] - comparison["current_routes"]
    comparison["route_reduction_pct"] = np.where(
        comparison["current_routes"] > 0,
        (comparison["current_routes"] - comparison["optimized_routes"]) / comparison["current_routes"],
        np.nan,
    )
    for metric in [
        "pair_mean_km",
        "nearest_mean_km",
        "radius_mean_km",
        "route_to_lot_centroid_mean_km",
        "route_size_cv",
        "separation_index_mean",
        "nn_path_haversine_km",
        "nn_path_geopandas_km",
    ]:
        comparison[f"{metric}_delta"] = (
            comparison[f"optimized_{metric}"] - comparison[f"current_{metric}"]
        )
    for metric in ["nn_path_haversine_km", "nn_path_geopandas_km"]:
        current_metric = comparison[f"current_{metric}"]
        comparison[f"{metric}_reduction_pct"] = np.where(
            current_metric.notna() & current_metric.ne(0),
            (current_metric - comparison[f"optimized_{metric}"]) / current_metric,
            np.nan,
        )
    return comparison.sort_values("lot").reset_index(drop=True)


def save_dashboard_snapshot(
    analysis: dict[str, Any],
    output_file: str | Path = DEFAULT_DASHBOARD_SNAPSHOT,
) -> Path:
    """Salva os dados já calculados que serão apenas lidos pelo Streamlit."""
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    temporary_file = output_file.with_name(f".{output_file.name}.tmp")
    payload = dict(analysis)
    # O caminho absoluto das fontes identifica o computador local e não é
    # necessário para o dashboard publicado. Mantemos apenas os metadados úteis.
    payload["source_info"] = [
        {key: value for key, value in item.items() if key != "path"}
        for item in analysis.get("source_info", [])
    ]
    payload["lot_comparison"] = build_lot_comparison(analysis)
    payload["snapshot_version"] = DASHBOARD_SNAPSHOT_VERSION
    payload["snapshot_generated_at"] = datetime.now()
    try:
        with gzip.open(temporary_file, "wb", compresslevel=3) as stream:
            pickle.dump(payload, stream, protocol=pickle.HIGHEST_PROTOCOL)
        temporary_file.replace(output_file)
    finally:
        temporary_file.unlink(missing_ok=True)
    return output_file.resolve()


def load_dashboard_snapshot(
    snapshot_file: str | Path = DEFAULT_DASHBOARD_SNAPSHOT,
) -> dict[str, Any]:
    """Carrega um snapshot local e confiável gerado por analise_rotas.ipynb."""
    snapshot_file = Path(snapshot_file)
    if not snapshot_file.exists():
        raise FileNotFoundError(
            f"Snapshot do dashboard não encontrado: {snapshot_file}. Execute Run All em analise_rotas.ipynb."
        )
    with gzip.open(snapshot_file, "rb") as stream:
        payload = pickle.load(stream)
    if not isinstance(payload, dict):
        raise ValueError("O snapshot do dashboard possui formato inválido.")
    if payload.get("snapshot_version") != DASHBOARD_SNAPSHOT_VERSION:
        raise ValueError(
            "O snapshot do dashboard está desatualizado. Execute Run All em analise_rotas.ipynb."
        )
    required = {
        "current", "optimized", "route_metrics", "lot_metrics", "coverage",
        "transitions", "stability", "fragmentation", "quality_summary",
        "quality_issues", "source_info", "lot_comparison",
    }
    missing = required.difference(payload)
    if missing:
        raise ValueError(
            "O snapshot do dashboard está incompleto: " + ", ".join(sorted(missing))
        )
    return payload


def source_signature(paths: Iterable[str | Path]) -> tuple[tuple[str, int, int], ...]:
    signature: list[tuple[str, int, int]] = []
    for raw_path in paths:
        path = Path(raw_path)
        stat = path.stat()
        signature.append((str(path.resolve()), int(stat.st_mtime_ns), int(stat.st_size)))
    return tuple(signature)


def aligned_route_colors(
    current_lot: pd.DataFrame,
    optimized_lot: pd.DataFrame,
    transitions_lot: pd.DataFrame,
) -> tuple[dict[str, list[int]], dict[str, list[int]]]:
    current_routes = sorted(current_lot["route"].astype(str).unique())
    optimized_routes = sorted(optimized_lot["route"].astype(str).unique())
    total_colors = max(len(current_routes) + len(optimized_routes), 1)

    def palette(index: int) -> list[int]:
        hue = (index * 0.618033988749895) % 1.0
        red, green, blue = colorsys.hsv_to_rgb(hue, 0.68, 0.90)
        return [int(red * 255), int(green * 255), int(blue * 255), 190]

    current_colors = {route: palette(index) for index, route in enumerate(current_routes)}
    optimized_colors: dict[str, list[int]] = {}

    if current_routes and optimized_routes and not transitions_lot.empty:
        current_index = {route: index for index, route in enumerate(current_routes)}
        optimized_index = {route: index for index, route in enumerate(optimized_routes)}
        matrix = np.zeros((len(current_routes), len(optimized_routes)), dtype=float)
        for _, row in transitions_lot.iterrows():
            current_route = str(row["current_route"])
            optimized_route = str(row["optimized_route"])
            if current_route in current_index and optimized_route in optimized_index:
                matrix[current_index[current_route], optimized_index[optimized_route]] = float(row["installations"])
        rows, columns = linear_sum_assignment(-matrix)
        for row_index, column_index in zip(rows, columns):
            if matrix[row_index, column_index] > 0:
                optimized_colors[optimized_routes[column_index]] = current_colors[current_routes[row_index]]

    next_color = len(current_routes)
    for route in optimized_routes:
        if route not in optimized_colors:
            optimized_colors[route] = palette(next_color)
            next_color += 1
    return current_colors, optimized_colors


def _excel_value(value: Any) -> Any:
    if value is pd.NA or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _write_dataframe(
    worksheet: Any,
    dataframe: pd.DataFrame,
    title: str,
    subtitle: str,
    start_row: int = 1,
    add_filter: bool = True,
) -> tuple[int, int]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title_row = start_row
    subtitle_row = start_row + 1
    header_row = start_row + 2
    first_data_row = start_row + 3
    last_column = max(len(dataframe.columns), 1)
    last_letter = get_column_letter(last_column)
    worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_column)
    worksheet.cell(title_row, 1, title)
    worksheet.cell(title_row, 1).font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    worksheet.cell(title_row, 1).fill = PatternFill("solid", fgColor="17365D")
    worksheet.cell(title_row, 1).alignment = Alignment(vertical="center")
    worksheet.row_dimensions[title_row].height = 26

    worksheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=last_column)
    worksheet.cell(subtitle_row, 1, subtitle)
    worksheet.cell(subtitle_row, 1).font = Font(name="Aptos", size=9, italic=True, color="44546A")
    worksheet.cell(subtitle_row, 1).alignment = Alignment(vertical="center")
    worksheet.row_dimensions[subtitle_row].height = 20

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_border = Border(bottom=Side(style="thin", color="8EA9C1"))
    for column_index, column in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(header_row, column_index, str(column))
        cell.font = Font(name="Aptos", size=10, bold=True, color="17365D")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 34

    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=first_data_row):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, column_index, _excel_value(value))
            cell.font = Font(name="Aptos", size=9, color="1F1F1F")
            cell.alignment = Alignment(vertical="center")

    last_data_row = first_data_row + max(len(dataframe) - 1, 0)
    if add_filter and len(dataframe):
        worksheet.auto_filter.ref = f"A{header_row}:{last_letter}{last_data_row}"
    worksheet.sheet_view.showGridLines = False

    sample = dataframe.head(1000)
    for column_index, column in enumerate(dataframe.columns, start=1):
        values = [str(column)] + [str(value) for value in sample[column].dropna().tolist()]
        width = min(max(max((len(value) for value in values), default=8) + 2, 10), 38)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
        column_name = str(column).lower()
        if len(dataframe):
            for row_index in range(first_data_row, last_data_row + 1):
                cell = worksheet.cell(row_index, column_index)
                if "%" in str(column) or "percentual" in column_name or "share" in column_name or column_name in {
                    "retenção de pares", "precisão de pares", "jaccard", "ari", "nmi"
                }:
                    cell.number_format = "0.0%"
                elif "km" in column_name:
                    cell.number_format = "0.000"
                elif "latitude" in column_name or "longitude" in column_name:
                    cell.number_format = "0.000000"
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000" if not float(cell.value).is_integer() else "#,##0"
    return header_row, last_data_row


def _route_export(route_metrics: pd.DataFrame) -> pd.DataFrame:
    columns = {
        "scenario": "Cenário",
        "lot": "Lote",
        "route": "UL",
        "installations": "Instalações",
        "centroid_latitude": "Latitude centroide",
        "centroid_longitude": "Longitude centroide",
        "pair_mean_km": "Pares média km",
        "pair_median_km": "Pares mediana km",
        "pair_max_km": "Pares máxima km",
        "pair_zero_pct": "Pares distância zero %",
        "nearest_mean_km": "Vizinho média km",
        "nearest_median_km": "Vizinho mediana km",
        "nearest_max_km": "Vizinho máxima km",
        "nearest_zero_pct": "Vizinho distância zero %",
        "radius_mean_km": "Instalação-centroide média km",
        "radius_median_km": "Instalação-centroide mediana km",
        "radius_max_km": "Instalação-centroide máxima km",
        "route_to_lot_centroid_km": "Centroide UL-lote km",
        "nearest_route_centroid_km": "Centroide UL mais próximo km",
        "separation_index": "Índice de separação",
    }
    return route_metrics[list(columns)].rename(columns=columns)


def _summary_export(analysis: dict[str, Any]) -> pd.DataFrame:
    comparison = build_lot_comparison(analysis)
    result = pd.DataFrame(
        {
            "Lote": comparison["lot"],
            "Status": comparison["status"],
            "Instalações atual": comparison["current_installations"],
            "Instalações otimizada": comparison["optimized_installations"],
            "Instalações comuns": comparison["common_installations"],
            "Cobertura atual %": comparison["current_coverage_pct"],
            "Rotas atual": comparison["current_routes"],
            "Rotas otimizada": comparison["optimized_routes"],
            "Delta rotas": comparison["route_delta"],
            "Redução de rotas %": comparison["route_reduction_pct"],
            "Pares média atual km": comparison["current_pair_mean_km"],
            "Pares média otimizada km": comparison["optimized_pair_mean_km"],
            "Delta pares média km": comparison["pair_mean_km_delta"],
            "Delta pares média %": np.where(
                comparison["current_pair_mean_km"].notna() & comparison["current_pair_mean_km"].ne(0),
                comparison["pair_mean_km_delta"] / comparison["current_pair_mean_km"],
                np.nan,
            ),
            "Vizinho média atual km": comparison["current_nearest_mean_km"],
            "Vizinho média otimizada km": comparison["optimized_nearest_mean_km"],
            "Delta vizinho média km": comparison["nearest_mean_km_delta"],
            "Raio médio atual km": comparison["current_radius_mean_km"],
            "Raio médio otimizada km": comparison["optimized_radius_mean_km"],
            "Delta raio médio km": comparison["radius_mean_km_delta"],
            "Centroide UL-lote atual km": comparison["current_route_to_lot_centroid_mean_km"],
            "Centroide UL-lote otimizada km": comparison["optimized_route_to_lot_centroid_mean_km"],
            "Delta centroide UL-lote km": comparison["route_to_lot_centroid_mean_km_delta"],
            "Retenção de pares": comparison["pair_retention"],
            "Precisão de pares": comparison["pair_precision"],
            "Jaccard": comparison["pair_jaccard"],
            "ARI": comparison["ari"],
            "NMI": comparison["nmi"],
        }
    )
    return result


def build_workbook(analysis: dict[str, Any]) -> Any:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    subtitle = f"Gerado em {analysis['generated_at']:%d/%m/%Y %H:%M} | Distâncias Haversine em km"

    summary = _summary_export(analysis)
    ws = workbook.create_sheet("Resumo_Lotes")
    header_row, last_row = _write_dataframe(ws, summary, "Comparativo de rotas por lote", subtitle)
    ws.freeze_panes = "C4"
    # Simple deltas remain auditable Excel formulas.
    formula_columns = {
        9: "=IF(AND(G{row}>0,H{row}>0),IFERROR(H{row}-G{row},\"\"),\"\")",
        10: "=IF(AND(G{row}>0,H{row}>0),IFERROR((G{row}-H{row})/G{row},\"\"),\"\")",
        13: "=IF(AND(G{row}>0,H{row}>0),IFERROR(L{row}-K{row},\"\"),\"\")",
        14: "=IF(AND(G{row}>0,H{row}>0),IFERROR(M{row}/K{row},\"\"),\"\")",
        17: "=IF(AND(G{row}>0,H{row}>0),IFERROR(P{row}-O{row},\"\"),\"\")",
        20: "=IF(AND(G{row}>0,H{row}>0),IFERROR(S{row}-R{row},\"\"),\"\")",
        23: "=IF(AND(G{row}>0,H{row}>0),IFERROR(V{row}-U{row},\"\"),\"\")",
    }
    for row in range(header_row + 1, last_row + 1):
        for column, formula in formula_columns.items():
            ws.cell(row, column, formula.format(row=row))
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    for column in [9, 13, 17, 20, 23]:
        letter = get_column_letter(column)
        ws.conditional_formatting.add(
            f"{letter}{header_row + 1}:{letter}{last_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"{letter}{header_row + 1}:{letter}{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill),
        )
    ws.conditional_formatting.add(
        f"B{header_row + 1}:B{last_row}",
        FormulaRule(formula=[f'$B{header_row + 1}="Comparável"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"X{header_row + 1}:Z{last_row}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
                       mid_color="FFEB84", end_type="max", end_color="63BE7B"),
    )

    ws = workbook.create_sheet("Indicadores_UL")
    _, last_route_row = _write_dataframe(
        ws,
        _route_export(analysis["route_metrics"]),
        "Indicadores espaciais por UL",
        subtitle,
    )
    ws.freeze_panes = "D4"

    transition_columns = {
        "lot": "Lote",
        "current_route": "UL atual",
        "optimized_route": "UL otimizada",
        "installations": "Instalações comuns",
        "current_route_installations": "Total da UL atual",
        "optimized_route_installations": "Total da UL otimizada",
        "share_current_route": "Participação na UL atual %",
        "share_optimized_route": "Participação na UL otimizada %",
        "retained_pairs_in_cell": "Pares mantidos na célula",
    }
    ws = workbook.create_sheet("Transicoes_UL")
    _write_dataframe(
        ws,
        analysis["transitions"][list(transition_columns)].rename(columns=transition_columns),
        "Transição de instalações entre ULs",
        "As ULs são relacionadas pelas instalações comuns, nunca pelo identificador.",
    )
    ws.freeze_panes = "D4"

    stability = analysis["stability"].rename(columns={
        "lot": "Lote",
        "common_installations": "Instalações comuns",
        "current_pairs": "Pares atuais",
        "optimized_pairs": "Pares otimizados",
        "retained_pairs": "Pares mantidos",
        "separated_pairs": "Pares separados",
        "newly_grouped_pairs": "Novos pares",
        "pair_retention": "Retenção de pares",
        "pair_precision": "Precisão de pares",
        "pair_jaccard": "Jaccard",
        "ari": "ARI",
        "nmi": "NMI",
    })
    fragmentation = analysis["fragmentation"].rename(columns={
        "lot": "Lote",
        "current_route": "UL atual",
        "common_installations": "Instalações comuns",
        "optimized_routes_received": "ULs otimizadas recebidas",
        "dominant_optimized_route": "UL otimizada dominante",
        "dominant_installations": "Instalações no grupo dominante",
        "dominant_share": "Participação dominante %",
        "fragmentation_index": "Índice de fragmentação",
        "separated_pairs": "Pares separados",
        "separated_pairs_pct": "Pares separados %",
    })
    ws = workbook.create_sheet("Estabilidade")
    _, stability_end = _write_dataframe(
        ws,
        stability,
        "Estabilidade dos agrupamentos por lote",
        subtitle,
        add_filter=False,
    )
    second_start = stability_end + 3
    _write_dataframe(
        ws,
        fragmentation,
        "Fragmentação das ULs atuais",
        "Quanto menor a fragmentação e o percentual de pares separados, maior a continuidade do grupo.",
        start_row=second_start,
    )
    ws.freeze_panes = "A4"

    coverage = analysis["coverage"].rename(columns={
        "lot": "Lote",
        "status": "Status",
        "current_installations": "Instalações atual",
        "optimized_installations": "Instalações otimizada",
        "common_installations": "Instalações comuns",
        "current_only_installations": "Somente atual",
        "optimized_only_installations": "Somente otimizada",
        "current_routes": "Rotas atual",
        "optimized_routes": "Rotas otimizada",
        "current_coverage_pct": "Cobertura atual %",
    })
    quality_summary = analysis["quality_summary"].rename(columns={
        "scenario": "Cenário", "metric": "Indicador", "value": "Valor", "note": "Observação"
    })
    quality_issues = analysis["quality_issues"].rename(columns={
        "scenario": "Cenário", "issue": "Ocorrência", "lot": "Lote",
        "installation": "Instalação", "route": "UL", "detail": "Detalhe"
    })
    ws = workbook.create_sheet("Cobertura_Qualidade")
    _, coverage_end = _write_dataframe(
        ws,
        coverage,
        "Cobertura e disponibilidade por lote",
        subtitle,
        add_filter=False,
    )
    _, quality_end = _write_dataframe(
        ws,
        quality_summary,
        "Resumo de qualidade das fontes",
        "Registros descartados ou ajustados antes dos indicadores.",
        start_row=coverage_end + 3,
        add_filter=False,
    )
    _write_dataframe(
        ws,
        quality_issues,
        "Ocorrências para conferência",
        "A listagem é limitada às primeiras 1.000 ocorrências de cada tipo e cenário.",
        start_row=quality_end + 3,
    )
    ws.freeze_panes = "A4"

    match = analysis["installation_match"].rename(columns={
        "lot": "Lote",
        "installation": "Instalação",
        "current_route": "UL atual",
        "optimized_route": "UL otimizada",
        "current_latitude": "Latitude atual",
        "current_longitude": "Longitude atual",
        "optimized_latitude": "Latitude otimizada",
        "optimized_longitude": "Longitude otimizada",
        "status": "Status",
        "coordinate_difference_m": "Diferença de coordenada m",
    })
    ws = workbook.create_sheet("Instalacoes_Match")
    _write_dataframe(
        ws,
        match,
        "Correspondência de instalações nos lotes comparáveis",
        "A chave de correspondência é Lote + Instalação; IDs de UL não são comparados diretamente.",
    )
    ws.freeze_panes = "C4"

    dictionary_rows = [
        ("Pares média/mediana/máxima", "Distância Haversine de todas as combinações únicas de duas instalações da mesma UL."),
        ("Vizinho média/mediana/máxima", "Para cada instalação, distância Haversine até a instalação mais próxima da mesma UL."),
        ("Instalação-centroide", "Distância entre cada instalação e a média de latitude/longitude das instalações de sua UL."),
        ("Centroide UL-lote", "Distância do centroide da UL ao centroide do lote no mesmo cenário."),
        ("Índice de separação", "Distância ao centroide de UL mais próximo dividida pelo raio médio da UL; valores maiores indicam melhor separação."),
        ("Retenção de pares", "Percentual dos pares que estavam juntos na UL atual e continuam juntos na otimizada."),
        ("Precisão de pares", "Percentual dos pares juntos na otimizada que também estavam juntos na atual."),
        ("Jaccard", "Interseção sobre união das relações de coagrupamento atual e otimizada."),
        ("ARI", "Adjusted Rand Index entre os dois agrupamentos, calculado somente nas instalações comuns do lote."),
        ("NMI", "Normalized Mutual Information entre os agrupamentos, calculada somente nas instalações comuns do lote."),
        ("Índice de fragmentação", "1 menos a soma dos quadrados das participações de uma UL atual nas ULs otimizadas."),
        ("Cobertura atual", "Instalações comuns divididas pelas instalações existentes no cenário atual do lote."),
        ("Escopo", "Indicadores espaciais usam o cenário completo disponível; estabilidade usa somente instalações comuns no mesmo lote."),
        ("Limitação", "Não há sequência de visitas ou rede viária; os indicadores medem estrutura espacial dos grupos, não percurso rodoviário."),
    ]
    dictionary = pd.DataFrame(dictionary_rows, columns=["Indicador", "Definição"])
    ws = workbook.create_sheet("Dicionario")
    _write_dataframe(ws, dictionary, "Dicionário de indicadores", subtitle)
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    for row in range(4, 4 + len(dictionary)):
        from openpyxl.styles import Alignment

        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 34

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass
    return workbook


def workbook_bytes(analysis: dict[str, Any]) -> bytes:
    buffer = BytesIO()
    workbook = build_workbook(analysis)
    workbook.save(buffer)
    return buffer.getvalue()


def export_workbook(
    analysis: dict[str, Any],
    output_file: str | Path = DEFAULT_OUTPUT_FILE,
) -> Path:
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(analysis)
    workbook.save(output_file)
    return output_file.resolve()


if __name__ == "__main__":
    result = run_analysis()
    output = export_workbook(result)
    comparison = build_lot_comparison(result)
    print(comparison[["lot", "status", "current_installations", "optimized_installations", "current_routes", "optimized_routes"]].to_string(index=False))
    print(f"\nArquivo gerado: {output}")
